#!/usr/bin/env python3
"""
biotech_screener_v1_clean.py

Excel-first biotech screener MVP.

What it does
------------
1) Builds a public-company universe from SEC ticker data.
2) Pulls SEC submissions metadata and XBRL company facts.
3) Pulls ClinicalTrials.gov study data by sponsor/company name.
4) Optionally pulls FDA Drugs@FDA approval/application data by sponsor name.
5) Writes a real .xlsx workbook.

Sheets written
--------------
- CompanySummary
- Universe
- TrialsRaw
- FDARaw
- Unmatched
- Errors
"""

from __future__ import annotations

import argparse
import re
import sys
import time
import traceback
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zipfile import ZipFile

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SEC_TICKERS_EXCHANGE_URL = "https://www.sec.gov/files/company_tickers_exchange.json"
SEC_SUBMISSIONS_URL = "https://data.sec.gov/submissions/CIK{cik}.json"
SEC_COMPANYFACTS_URL = "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
CTG_LEGACY_STUDY_FIELDS_URL = "https://clinicaltrials.gov/api/query/study_fields"
DRUGS_FDA_ZIP_URL = "https://www.fda.gov/media/89850/download"

DEFAULT_HEADERS = {
    "User-Agent": "FerdiBiotechScreener/0.1 (educational research tool; replace-with-your-email@example.com)",
    "Accept": "application/json, text/plain, */*",
}

BIOTECH_KEYWORDS = [
    "biotech", "biotechnology", "biopharma", "biopharmaceutical", "pharmaceutical",
    "therapeutics", "oncology", "rare disease", "clinical-stage", "clinical stage",
    "drug discovery", "pipeline", "immunotherapy", "cell therapy", "gene therapy",
    "precision medicine", "antibody", "rna", "mrna", "crispr", "small molecule",
]

NAME_STOPWORDS = {
    "inc", "inc.", "corp", "corp.", "corporation", "company", "co", "co.",
    "holdings", "holding", "group", "ltd", "ltd.", "limited", "plc", "ag",
    "sa", "nv", "se", "the", "biosciences", "bioscience", "biologics"
}

XBRL_CASH_TAGS = [
    ("us-gaap", "CashAndCashEquivalentsAtCarryingValue"),
    ("us-gaap", "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"),
]
XBRL_ASSET_TAGS = [("us-gaap", "Assets")]
XBRL_OCF_TAGS = [
    ("us-gaap", "NetCashProvidedByUsedInOperatingActivities"),
    ("us-gaap", "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations"),
]
XBRL_NET_INCOME_TAGS = [("us-gaap", "NetIncomeLoss")]

PHASE_ORDER = {
    "EARLY_PHASE1": 1,
    "PHASE1": 2,
    "PHASE1|PHASE2": 3,
    "PHASE2": 4,
    "PHASE2|PHASE3": 5,
    "PHASE3": 6,
    "PHASE4": 7,
    "NA": 0,
    "N/A": 0,
    "": 0,
}

RAW_TRIAL_FIELDS = [
    "NCTId",
    "BriefTitle",
    "Condition",
    "InterventionName",
    "InterventionType",
    "LeadSponsorName",
    "CollaboratorName",
    "Phase",
    "OverallStatus",
    "PrimaryCompletionDate",
    "CompletionDate",
    "StudyFirstSubmitDate",
    "LastUpdatePostDate",
]


@dataclass
class Company:
    ticker: str
    cik: str
    company_name: str
    exchange: str = ""
    sic: str = ""
    sic_description: str = ""
    aliases: List[str] = field(default_factory=list)
    biotech_flag: bool = False
    biotech_reason: str = ""
    facts: Dict[str, Any] = field(default_factory=dict)


@dataclass
class TrialRecord:
    ticker: str
    company_name: str
    sponsor_query: str
    nct_id: str = ""
    brief_title: str = ""
    condition: str = ""
    intervention_name: str = ""
    intervention_type: str = ""
    lead_sponsor_name: str = ""
    collaborator_name: str = ""
    phase: str = ""
    overall_status: str = ""
    primary_completion_date: str = ""
    completion_date: str = ""
    study_first_submit_date: str = ""
    last_update_post_date: str = ""


@dataclass
class FDARow:
    ticker: str
    company_name: str
    sponsor_name: str
    appl_no: str = ""
    appl_type: str = ""
    submission_type: str = ""
    submission_status_date: str = ""
    product_no: str = ""
    drug_name: str = ""
    active_ingredient: str = ""
    marketing_status: str = ""


def log(msg: str) -> None:
    print(msg, flush=True)


def parse_date_safe(value: Any) -> Optional[date]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y"):
        try:
            if fmt == "%Y-%m":
                dt = datetime.strptime(s[:7], fmt)
                return date(dt.year, dt.month, 1)
            if fmt == "%Y-%m-%d":
                dt = datetime.strptime(s[:10], fmt)
                return dt.date()
            if fmt == "%Y-%m-%d %H:%M:%S":
                dt = datetime.strptime(s[:19], fmt)
                return dt.date()
            if fmt == "%m/%d/%Y":
                dt = datetime.strptime(s[:10], fmt)
                return dt.date()
        except Exception:
            pass

    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except Exception:
        return None


def days_until(value: Any) -> Optional[int]:
    d = parse_date_safe(value)
    if d is None:
        return None
    return (d - date.today()).days


def normalize_name(name: str) -> str:
    s = (name or "").lower()
    s = re.sub(r"[^a-z0-9\s&-]", " ", s)
    tokens = [t for t in re.split(r"\s+", s) if t]
    tokens = [t for t in tokens if t not in NAME_STOPWORDS]
    return " ".join(tokens).strip()


def possible_aliases(company_name: str) -> List[str]:
    raw = company_name.strip()
    norm = normalize_name(raw)

    aliases = {raw}
    if norm:
        aliases.add(norm)

    # remove common legal suffixes and location fragments
    cleaned = re.sub(r"\s*/\s*[A-Z]{2,3}$", "", raw).strip()
    cleaned = re.sub(
        r"\b(incorporated|inc\.?|corp\.?|corporation|company|co\.?|ltd\.?|limited|plc|ag|sa|nv|se)\b",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ,/")
    if cleaned:
        aliases.add(cleaned)
        aliases.add(cleaned.lower())

    norm_clean = normalize_name(cleaned)
    if norm_clean:
        aliases.add(norm_clean)

        tokens = norm_clean.split()
        if len(tokens) >= 2:
            aliases.add(" ".join(tokens[:2]))
        if len(tokens) >= 3:
            aliases.add(" ".join(tokens[:3]))
        if len(tokens) >= 1:
            aliases.add(tokens[0])

    # a few manual-style biotech friendly variants
    if "pharmaceuticals" in norm:
        aliases.add(norm.replace("pharmaceuticals", "").strip())
        aliases.add(norm.replace("pharmaceuticals", "pharma").strip())
    if "therapeutics" in norm:
        aliases.add(norm.replace("therapeutics", "").strip())
    if "moderna inc" in norm or "moderna" in norm:
        aliases.add("ModernaTX")
        aliases.add("ModernaTX, Inc.")
    if "vertex pharmaceuticals" in norm:
        aliases.add("Vertex Pharmaceuticals Incorporated")
    if "crispr therapeutics" in norm:
        aliases.add("CRISPR Therapeutics")

    return [a for a in sorted(aliases, key=len, reverse=True) if a]


def safe_float(x: Any) -> Optional[float]:
    try:
        if x is None or x == "":
            return None
        return float(x)
    except Exception:
        return None


def latest_fact_value(
    facts_json: Dict[str, Any],
    taxonomy: str,
    tag: str,
    unit_preference: Optional[str] = None,
) -> Tuple[Optional[float], Optional[str], Optional[str]]:
    facts = facts_json.get("facts", {}).get(taxonomy, {}).get(tag, {})
    units = facts.get("units", {})
    if not units:
        return None, None, None

    candidates = []
    unit_keys = [unit_preference] if unit_preference and unit_preference in units else list(units.keys())
    for unit in unit_keys:
        if unit not in units:
            continue
        for item in units[unit]:
            val = safe_float(item.get("val"))
            end = item.get("end")
            form = item.get("form")
            fy = item.get("fy")
            fp = item.get("fp")
            if val is None or not end:
                continue
            parsed = parse_date_safe(end)
            if parsed is None:
                continue
            candidates.append((parsed, val, end, form, fy, fp))

    if not candidates:
        return None, None, None

    candidates.sort(key=lambda x: x[0], reverse=True)
    _, val, end, form, _, _ = candidates[0]
    return val, end, form


def first_available_fact(
    facts_json: Dict[str, Any],
    choices: List[Tuple[str, str]],
    unit_preference: Optional[str] = "USD",
) -> Tuple[Optional[float], Optional[str], Optional[str], Optional[str]]:
    for taxonomy, tag in choices:
        val, end, form = latest_fact_value(facts_json, taxonomy, tag, unit_preference=unit_preference)
        if val is not None:
            return val, end, form, f"{taxonomy}:{tag}"
    return None, None, None, None


def request_json(
    url: str,
    session: requests.Session,
    params: Optional[Dict[str, Any]] = None,
    sleep_sec: float = 0.15,
) -> Dict[str, Any]:
    r = session.get(url, params=params, timeout=60)
    r.raise_for_status()
    time.sleep(sleep_sec)
    return r.json()


def request_bytes(url: str, session: requests.Session) -> bytes:
    r = session.get(url, timeout=120)
    r.raise_for_status()
    return r.content


def load_sec_ticker_universe(session: requests.Session) -> List[Company]:
    data = request_json(SEC_TICKERS_EXCHANGE_URL, session)
    companies: List[Company] = []

    if not (isinstance(data, dict) and "data" in data):
        raise ValueError("Unexpected SEC tickers payload format")

    rows = data["data"]
    fields = data.get("fields", [])
    idx = {name: i for i, name in enumerate(fields)}

    for row in rows:
        try:
            ticker = str(row[idx["ticker"]]).strip()
            cik_raw = str(row[idx["cik"]]).strip()
            company_name = str(row[idx["name"]]).strip()
            exchange = str(row[idx["exchange"]]).strip() if "exchange" in idx else ""
        except Exception:
            continue

        if not ticker or not cik_raw or not company_name:
            continue

        companies.append(
            Company(
                ticker=ticker.upper(),
                cik=str(cik_raw).zfill(10),
                company_name=company_name,
                exchange=exchange,
                aliases=possible_aliases(company_name),
            )
        )

    return companies


def enrich_submissions_metadata(company: Company, session: requests.Session) -> None:
    data = request_json(SEC_SUBMISSIONS_URL.format(cik=company.cik), session)
    company.sic = str(data.get("sic", "") or "")
    company.sic_description = str(data.get("sicDescription", "") or "")

    former_names = data.get("formerNames") or []
    former_name_text = " ".join(
        item.get("name", "") if isinstance(item, dict) else str(item)
        for item in former_names
    )

    haystack = " ".join([
        company.company_name,
        company.sic_description,
        former_name_text,
    ]).lower()

    matched = [kw for kw in BIOTECH_KEYWORDS if kw in haystack]
    if matched:
        company.biotech_flag = True
        company.biotech_reason = f"keyword:{matched[0]}"
    elif company.sic_description and any(x in company.sic_description.lower() for x in ["pharmaceutical", "biological", "medical", "drug"]):
        company.biotech_flag = True
        company.biotech_reason = "sic_description"

    filings = data.get("filings", {}).get("recent", {})
    forms = filings.get("form", []) or []
    filing_dates = filings.get("filingDate", []) or []

    company.facts["is_alive"] = False
    company.facts["alive_reason"] = "no recent SEC filings"

    latest_relevant_date = None
    latest_relevant_form = ""
    for form, filing_date in zip(forms, filing_dates):
        if form in {"10-K", "10-Q", "20-F", "6-K", "8-K", "S-1", "F-1"}:
            d = parse_date_safe(filing_date)
            if d and (latest_relevant_date is None or d > latest_relevant_date):
                latest_relevant_date = d
                latest_relevant_form = form

    if latest_relevant_date is not None:
        age_days = (date.today() - latest_relevant_date).days
        company.facts["latest_filing_date"] = latest_relevant_date.isoformat()
        company.facts["latest_filing_form"] = latest_relevant_form
        company.facts["days_since_latest_filing"] = age_days
        if age_days <= 730:
            company.facts["is_alive"] = True
            company.facts["alive_reason"] = f"recent filing {latest_relevant_form} on {latest_relevant_date.isoformat()}"
        else:
            company.facts["alive_reason"] = f"stale filing history; last relevant filing {latest_relevant_form} on {latest_relevant_date.isoformat()}"

def load_companyfacts(company: Company, session: requests.Session) -> None:
    facts_json = request_json(SEC_COMPANYFACTS_URL.format(cik=company.cik), session)
    cash, cash_end, cash_form, cash_tag = first_available_fact(facts_json, XBRL_CASH_TAGS)
    assets, assets_end, assets_form, assets_tag = first_available_fact(facts_json, XBRL_ASSET_TAGS)
    ocf, ocf_end, ocf_form, ocf_tag = first_available_fact(facts_json, XBRL_OCF_TAGS)
    net_income, ni_end, ni_form, ni_tag = first_available_fact(facts_json, XBRL_NET_INCOME_TAGS)

    burn_annual = None
    if ocf is not None and ocf < 0:
        burn_annual = abs(ocf)
    elif net_income is not None and net_income < 0:
        burn_annual = abs(net_income)

    runway_months = None
    if cash is not None and burn_annual and burn_annual > 0:
        runway_months = (cash / burn_annual) * 12.0

    company.facts = {
        "cash": cash,
        "cash_end": cash_end,
        "cash_form": cash_form,
        "cash_tag": cash_tag,
        "assets": assets,
        "assets_end": assets_end,
        "assets_form": assets_form,
        "assets_tag": assets_tag,
        "operating_cash_flow": ocf,
        "ocf_end": ocf_end,
        "ocf_form": ocf_form,
        "ocf_tag": ocf_tag,
        "net_income": net_income,
        "net_income_end": ni_end,
        "net_income_form": ni_form,
        "net_income_tag": ni_tag,
        "estimated_annual_burn": burn_annual,
        "estimated_runway_months": runway_months,
    }


def ctg_fetch_studies_for_sponsor(
    sponsor_query: str,
    session: requests.Session,
    max_rank: int = 100,
) -> List[TrialRecord]:
    # Broad search first; we will post-filter in Python.
    params = {
        "expr": sponsor_query,
        "fields": ",".join(RAW_TRIAL_FIELDS),
        "min_rnk": 1,
        "max_rnk": max_rank,
        "fmt": "json",
    }

    data = request_json(CTG_LEGACY_STUDY_FIELDS_URL, session, params=params, sleep_sec=0.0)
    out: List[TrialRecord] = []

    study_fields = data.get("StudyFieldsResponse", {}).get("StudyFields", [])
    for row in study_fields:
        def first(field: str) -> str:
            v = row.get(field, [])
            if isinstance(v, list):
                return "; ".join(str(x) for x in v if x is not None)
            return str(v or "")

        out.append(
            TrialRecord(
                ticker="",
                company_name="",
                sponsor_query=sponsor_query,
                nct_id=first("NCTId"),
                brief_title=first("BriefTitle"),
                condition=first("Condition"),
                intervention_name=first("InterventionName"),
                intervention_type=first("InterventionType"),
                lead_sponsor_name=first("LeadSponsorName"),
                collaborator_name=first("CollaboratorName"),
                phase=first("Phase"),
                overall_status=first("OverallStatus"),
                primary_completion_date=first("PrimaryCompletionDate"),
                completion_date=first("CompletionDate"),
                study_first_submit_date=first("StudyFirstSubmitDate"),
                last_update_post_date=first("LastUpdatePostDate"),
            )
        )

    return out

    study_fields = data.get("StudyFieldsResponse", {}).get("StudyFields", [])
    for row in study_fields:
        def first(field: str) -> str:
            v = row.get(field, [])
            if isinstance(v, list):
                return "; ".join(str(x) for x in v if x is not None)
            return str(v or "")

        out.append(
            TrialRecord(
                ticker="",
                company_name="",
                sponsor_query=sponsor_query,
                nct_id=first("NCTId"),
                brief_title=first("BriefTitle"),
                condition=first("Condition"),
                intervention_name=first("InterventionName"),
                intervention_type=first("InterventionType"),
                lead_sponsor_name=first("LeadSponsorName"),
                collaborator_name=first("CollaboratorName"),
                phase=first("Phase"),
                overall_status=first("OverallStatus"),
                primary_completion_date=first("PrimaryCompletionDate"),
                completion_date=first("CompletionDate"),
                study_first_submit_date=first("StudyFirstSubmitDate"),
                last_update_post_date=first("LastUpdatePostDate"),
            )
        )

    return out

def token_set(text: str) -> set[str]:
    return {t for t in normalize_name(text).split() if t}

def sponsor_relevance_score(company: Company, alias: str, trial: TrialRecord) -> int:
    company_tokens = token_set(company.company_name)
    alias_tokens = token_set(alias)
    sponsor_tokens = token_set(trial.lead_sponsor_name)
    collab_tokens = token_set(trial.collaborator_name)

    overlap_lead_company = len(company_tokens & sponsor_tokens)
    overlap_lead_alias = len(alias_tokens & sponsor_tokens)
    overlap_collab_company = len(company_tokens & collab_tokens)
    overlap_collab_alias = len(alias_tokens & collab_tokens)

    score = max(
        overlap_lead_company * 3,
        overlap_lead_alias * 2,
        overlap_collab_company * 2,
        overlap_collab_alias,
    )

    sponsor_text = f"{trial.lead_sponsor_name} {trial.collaborator_name}".lower()
    norm_company = normalize_name(company.company_name)
    norm_alias = normalize_name(alias)

    if norm_company and norm_company in sponsor_text:
        score += 5
    if norm_alias and norm_alias in sponsor_text:
        score += 3

    return score

def filter_trials_for_company(company: Company, alias: str, trials: List[TrialRecord]) -> List[TrialRecord]:
    filtered = []
    for t in trials:
        score = sponsor_relevance_score(company, alias, t)
        if score >= 2:
            filtered.append(t)

    # deduplicate by NCT ID
    dedup = {}
    for t in filtered:
        dedup[t.nct_id or f"{t.brief_title}|{t.lead_sponsor_name}"] = t
    return list(dedup.values())


def choose_best_alias_trial_match(
    company: Company,
    session: requests.Session,
    max_trials_per_alias: int,
) -> Tuple[str, List[TrialRecord]]:
    best_alias = company.company_name
    best_trials: List[TrialRecord] = []
    best_score = -1

    for alias in company.aliases:
        try:
            raw_trials = ctg_fetch_studies_for_sponsor(alias, session, max_rank=max_trials_per_alias)
            trials = filter_trials_for_company(company, alias, raw_trials)
        except Exception:
            continue

        score = len(trials)
        if score > best_score:
            best_score = score
            best_alias = alias
            best_trials = trials

    for t in best_trials:
        t.ticker = company.ticker
        t.company_name = company.company_name

    return best_alias, best_trials


def load_drugsfda_tables(session: requests.Session) -> Dict[str, List[Dict[str, str]]]:
    blob = request_bytes(DRUGS_FDA_ZIP_URL, session)
    tables: Dict[str, List[Dict[str, str]]] = {}

    with ZipFile(BytesIO(blob)) as zf:
        names = {name.lower(): name for name in zf.namelist()}

        def read_tsv(possible_names: List[str]) -> List[Dict[str, str]]:
            for candidate in possible_names:
                lower = candidate.lower()
                if lower in names:
                    with zf.open(names[lower]) as fh:
                        content = fh.read().decode("utf-8", errors="replace").splitlines()
                    if not content:
                        return []
                    header = content[0].split("\t")
                    rows = []
                    for line in content[1:]:
                        parts = line.split("\t")
                        row = {header[i]: parts[i] if i < len(parts) else "" for i in range(len(header))}
                        rows.append(row)
                    return rows
            return []

        tables["applications"] = read_tsv(["Applications.txt"])
        tables["products"] = read_tsv(["Products.txt"])
        tables["submissions"] = read_tsv(["Submissions.txt"])
        tables["marketing_status"] = read_tsv(["MarketingStatus.txt"])
        tables["marketing_status_lookup"] = read_tsv(["MarketingStatus_Lookup.txt"])

    return tables


def sponsor_matches(company: Company, sponsor_name: str) -> bool:
    a = normalize_name(company.company_name)
    b = normalize_name(sponsor_name)
    if not a or not b:
        return False
    if a == b or a in b or b in a:
        return True
    a_set = set(a.split())
    b_set = set(b.split())
    overlap = len(a_set & b_set)
    return overlap >= 2 or (overlap >= 1 and min(len(a_set), len(b_set)) == 1)


def build_fda_rows_for_company(company: Company, fda_tables: Dict[str, List[Dict[str, str]]]) -> List[FDARow]:
    apps = fda_tables.get("applications", [])
    prods = fda_tables.get("products", [])
    subs = fda_tables.get("submissions", [])
    mstats = fda_tables.get("marketing_status", [])
    mlookup = {
        row.get("MarketingStatusID", ""): row.get("MarketingStatusDescription", "")
        for row in fda_tables.get("marketing_status_lookup", [])
    }

    products_by_appl: Dict[str, List[Dict[str, str]]] = {}
    for p in prods:
        products_by_appl.setdefault(p.get("ApplNo", ""), []).append(p)

    submissions_by_appl: Dict[str, List[Dict[str, str]]] = {}
    for s in subs:
        submissions_by_appl.setdefault(s.get("ApplNo", ""), []).append(s)

    marketing_by_key: Dict[Tuple[str, str], str] = {}
    for m in mstats:
        marketing_by_key[(m.get("ApplNo", ""), m.get("ProductNo", ""))] = mlookup.get(m.get("MarketingStatusID", ""), "")

    rows: List[FDARow] = []
    for app in apps:
        sponsor = app.get("SponsorName", "") or ""
        if not sponsor_matches(company, sponsor):
            continue

        appl_no = app.get("ApplNo", "") or ""
        appl_type = app.get("ApplType", "") or ""
        these_products = products_by_appl.get(appl_no, []) or [{}]
        these_subs = submissions_by_appl.get(appl_no, []) or [{}]

        latest_sub = None
        latest_sub_date = None
        for s in these_subs:
            d = parse_date_safe(s.get("SubmissionStatusDate"))
            if d and (latest_sub_date is None or d > latest_sub_date):
                latest_sub_date = d
                latest_sub = s

        for p in these_products:
            product_no = p.get("ProductNo", "") or ""
            rows.append(
                FDARow(
                    ticker=company.ticker,
                    company_name=company.company_name,
                    sponsor_name=sponsor,
                    appl_no=appl_no,
                    appl_type=appl_type,
                    submission_type=(latest_sub or {}).get("SubmissionType", "") if latest_sub else "",
                    submission_status_date=(latest_sub or {}).get("SubmissionStatusDate", "") if latest_sub else "",
                    product_no=product_no,
                    drug_name=p.get("DrugName", "") or "",
                    active_ingredient=p.get("ActiveIngredient", "") or "",
                    marketing_status=marketing_by_key.get((appl_no, product_no), ""),
                )
            )

    return rows


def phase_rank_value(phase_text: str) -> int:
    raw = (phase_text or "").upper().replace(" ", "")
    return PHASE_ORDER.get(raw, 0)


def summarize_trials(trials: List[TrialRecord]) -> Dict[str, Any]:
    if not trials:
        return {
            "study_count": 0,
            "unique_assets": 0,
            "highest_phase": "",
            "phase_distribution": "",
            "key_indications_top5": "",
            "lead_sponsor_count": 0,
            "collaborator_count": 0,
            "next_primary_completion_days": None,
            "next_completion_days": None,
            "catalysts_3m": 0,
            "catalysts_6m": 0,
            "catalysts_12m": 0,
            "late_stage_trials": 0,
            "terminated_withdrawn_suspended": 0,
            "single_asset_flag": True,
            "single_indication_flag": True,
        }

    unique_assets = set()
    indication_counter: Counter[str] = Counter()
    lead_sponsors = set()
    collaborators = set()
    phase_counter: Counter[str] = Counter()
    primary_days: List[int] = []
    completion_days: List[int] = []
    late_stage = 0
    risky_status = 0

    for t in trials:
        if t.intervention_name:
            for item in t.intervention_name.split(";"):
                item = item.strip()
                if item:
                    unique_assets.add(item)

        if t.condition:
            for cond in t.condition.split(";"):
                cond = cond.strip()
                if cond:
                    indication_counter[cond] += 1

        if t.lead_sponsor_name:
            lead_sponsors.add(t.lead_sponsor_name)

        if t.collaborator_name:
            for collab in t.collaborator_name.split(";"):
                collab = collab.strip()
                if collab:
                    collaborators.add(collab)

        phase = t.phase.strip()
        phase_counter[phase] += 1
        if phase_rank_value(phase) >= PHASE_ORDER["PHASE3"]:
            late_stage += 1

        status = (t.overall_status or "").lower()
        if any(x in status for x in ["terminated", "withdrawn", "suspended"]):
            risky_status += 1

        dp = days_until(t.primary_completion_date)
        if dp is not None and dp >= 0:
            primary_days.append(dp)

        dc = days_until(t.completion_date)
        if dc is not None and dc >= 0:
            completion_days.append(dc)

    highest_phase = max(phase_counter.keys(), key=phase_rank_value) if phase_counter else ""
    top_indications = ", ".join([k for k, _ in indication_counter.most_common(5)])
    phase_distribution = "; ".join([f"{k}:{v}" for k, v in phase_counter.most_common()])

    def count_within(arr: List[int], n: int) -> int:
        return sum(1 for x in arr if x <= n)

    return {
        "study_count": len(trials),
        "unique_assets": len(unique_assets),
        "highest_phase": highest_phase,
        "phase_distribution": phase_distribution,
        "key_indications_top5": top_indications,
        "lead_sponsor_count": len(lead_sponsors),
        "collaborator_count": len(collaborators),
        "next_primary_completion_days": min(primary_days) if primary_days else None,
        "next_completion_days": min(completion_days) if completion_days else None,
        "catalysts_3m": count_within(primary_days, 90),
        "catalysts_6m": count_within(primary_days, 180),
        "catalysts_12m": count_within(primary_days, 365),
        "late_stage_trials": late_stage,
        "terminated_withdrawn_suspended": risky_status,
        "single_asset_flag": len(unique_assets) <= 1,
        "single_indication_flag": len(indication_counter) <= 1,
    }


def summarize_fda(rows: List[FDARow]) -> Dict[str, Any]:
    if not rows:
        return {
            "approved_products_count": 0,
            "fda_latest_submission_date": "",
            "fda_products_top5": "",
        }

    products = [r.drug_name for r in rows if r.drug_name]
    latest_date = max(
        (parse_date_safe(r.submission_status_date) for r in rows if r.submission_status_date),
        default=None,
    )

    return {
        "approved_products_count": len(set(products)),
        "fda_latest_submission_date": latest_date.isoformat() if latest_date else "",
        "fda_products_top5": ", ".join(list(dict.fromkeys(products))[:5]),
    }


def autosize_worksheet(ws, max_width: int = 50) -> None:
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), max_width)


def write_sheet(ws, rows: List[Dict[str, Any]], title_bold: bool = True) -> None:
    if not rows:
        ws.append(["No rows"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    if title_bold:
        for cell in ws[1]:
            cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    autosize_worksheet(ws)


def load_alias_overrides(path: Optional[Path]) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    if not path or not path.exists():
        return out
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        if line.lower().startswith("ticker,"):
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 2:
            continue
        out.setdefault(parts[0].upper(), []).append(parts[1])
    return out


def run(args: argparse.Namespace) -> None:
    session = requests.Session()
    session.headers.update(DEFAULT_HEADERS)

    alias_overrides = load_alias_overrides(Path(args.alias_overrides) if args.alias_overrides else None)

    errors: List[Dict[str, str]] = []
    unmatched: List[Dict[str, str]] = []

    log("Loading SEC ticker universe...")
    universe = load_sec_ticker_universe(session)
    log(f"Loaded {len(universe):,} SEC tickers.")

    if args.tickers:
        wanted = {x.strip().upper() for x in args.tickers.split(",") if x.strip()}
        universe = [c for c in universe if c.ticker in wanted]
    elif args.limit:
        universe = universe[: args.limit]

    log("Applying Gate 1 (alive) and Gate 2 (biotech)...")
    filtered: List[Company] = []
    for i, company in enumerate(universe, start=1):
        try:
            enrich_submissions_metadata(company, session)
            if company.ticker in alias_overrides:
                company.aliases = list(dict.fromkeys(company.aliases + alias_overrides[company.ticker]))

            is_alive = bool(company.facts.get("is_alive", False))
            if not is_alive:
                unmatched.append({
                    "ticker": company.ticker,
                    "company_name": company.company_name,
                    "reason": "gate_1_fail_not_alive",
                    "aliases_tried": " | ".join(company.aliases),
                })
                continue

            if not company.biotech_flag:
                unmatched.append({
                    "ticker": company.ticker,
                    "company_name": company.company_name,
                    "reason": "gate_2_fail_not_biotech",
                    "aliases_tried": " | ".join(company.aliases),
                })
                continue

            filtered.append(company)
        except Exception as e:
            errors.append({
                "ticker": company.ticker,
                "company_name": company.company_name,
                "stage": "submissions",
                "error": str(e),
            })

        if i % 50 == 0:
            log(f"  gate 1/2 processed: {i}/{len(universe)}")

    universe = filtered
    log(f"Universe after Gate 1 and Gate 2: {len(universe):,}")

    fda_tables: Dict[str, List[Dict[str, str]]] = {}
    if args.include_fda:
        try:
            log("Loading Drugs@FDA tables...")
            fda_tables = load_drugsfda_tables(session)
        except Exception as e:
            errors.append({"ticker": "", "company_name": "", "stage": "drugsfda_load", "error": str(e)})

    summary_rows: List[Dict[str, Any]] = []
    universe_rows: List[Dict[str, Any]] = []
    trials_raw_rows: List[Dict[str, Any]] = []
    fda_raw_rows: List[Dict[str, Any]] = []

    log("Applying Gate 3 (has trial data) and running enrichment...")
    for i, company in enumerate(universe, start=1):
        try:
            load_companyfacts(company, session)
        except Exception as e:
            errors.append({
                "ticker": company.ticker,
                "company_name": company.company_name,
                "stage": "companyfacts",
                "error": str(e),
            })

        matched_alias = ""
        trial_records: List[TrialRecord] = []
        try:
            matched_alias, trial_records = choose_best_alias_trial_match(company, session, args.max_trials_per_alias)
        except Exception as e:
            errors.append({
                "ticker": company.ticker,
                "company_name": company.company_name,
                "stage": "clinicaltrials",
                "error": str(e),
            })

        if not trial_records:
            unmatched.append({
                "ticker": company.ticker,
                "company_name": company.company_name,
                "reason": "gate_3_fail_no_trial_data",
                "aliases_tried": " | ".join(company.aliases),
            })
            continue

        fda_rows_for_company: List[FDARow] = []
        if fda_tables:
            try:
                fda_rows_for_company = build_fda_rows_for_company(company, fda_tables)
            except Exception as e:
                errors.append({
                    "ticker": company.ticker,
                    "company_name": company.company_name,
                    "stage": "drugsfda_match",
                    "error": str(e),
                })

        trial_summary = summarize_trials(trial_records)
        fda_summary = summarize_fda(fda_rows_for_company)
        facts = company.facts

        universe_rows.append({
            "ticker": company.ticker,
            "cik": company.cik,
            "company_name": company.company_name,
            "exchange": company.exchange,
            "sic": company.sic,
            "sic_description": company.sic_description,
            "biotech_flag": company.biotech_flag,
            "biotech_reason": company.biotech_reason,
            "is_alive": facts.get("is_alive"),
            "alive_reason": facts.get("alive_reason"),
            "latest_filing_form": facts.get("latest_filing_form"),
            "latest_filing_date": facts.get("latest_filing_date"),
            "days_since_latest_filing": facts.get("days_since_latest_filing"),
            "aliases": " | ".join(company.aliases),
        })

        summary_rows.append({
            "ticker": company.ticker,
            "company_name": company.company_name,
            "matched_sponsor_alias": matched_alias,
            "study_count": trial_summary["study_count"],
            "unique_assets": trial_summary["unique_assets"],
            "highest_phase": trial_summary["highest_phase"],
            "phase_distribution": trial_summary["phase_distribution"],
            "key_indications_top5": trial_summary["key_indications_top5"],
            "lead_sponsor_count": trial_summary["lead_sponsor_count"],
            "collaborator_count": trial_summary["collaborator_count"],
            "next_primary_completion_days": trial_summary["next_primary_completion_days"],
            "next_completion_days": trial_summary["next_completion_days"],
            "catalysts_3m": trial_summary["catalysts_3m"],
            "catalysts_6m": trial_summary["catalysts_6m"],
            "catalysts_12m": trial_summary["catalysts_12m"],
            "late_stage_trials": trial_summary["late_stage_trials"],
            "cash": facts.get("cash"),
            "cash_end": facts.get("cash_end"),
            "cash_form": facts.get("cash_form"),
            "cash_tag": facts.get("cash_tag"),
            "assets": facts.get("assets"),
            "assets_end": facts.get("assets_end"),
            "operating_cash_flow": facts.get("operating_cash_flow"),
            "ocf_end": facts.get("ocf_end"),
            "net_income": facts.get("net_income"),
            "net_income_end": facts.get("net_income_end"),
            "estimated_annual_burn": facts.get("estimated_annual_burn"),
            "estimated_runway_months": facts.get("estimated_runway_months"),
            "terminated_withdrawn_suspended": trial_summary["terminated_withdrawn_suspended"],
            "single_asset_flag": trial_summary["single_asset_flag"],
            "single_indication_flag": trial_summary["single_indication_flag"],
            "approved_products_count": fda_summary["approved_products_count"],
            "fda_latest_submission_date": fda_summary["fda_latest_submission_date"],
            "fda_products_top5": fda_summary["fda_products_top5"],
        })

        for t in trial_records:
            trials_raw_rows.append({
                "ticker": t.ticker,
                "company_name": t.company_name,
                "sponsor_query": t.sponsor_query,
                "nct_id": t.nct_id,
                "brief_title": t.brief_title,
                "condition": t.condition,
                "intervention_name": t.intervention_name,
                "intervention_type": t.intervention_type,
                "lead_sponsor_name": t.lead_sponsor_name,
                "collaborator_name": t.collaborator_name,
                "phase": t.phase,
                "overall_status": t.overall_status,
                "primary_completion_date": t.primary_completion_date,
                "completion_date": t.completion_date,
                "study_first_submit_date": t.study_first_submit_date,
                "last_update_post_date": t.last_update_post_date,
            })

        for row in fda_rows_for_company:
            fda_raw_rows.append({
                "ticker": row.ticker,
                "company_name": row.company_name,
                "sponsor_name": row.sponsor_name,
                "appl_no": row.appl_no,
                "appl_type": row.appl_type,
                "submission_type": row.submission_type,
                "submission_status_date": row.submission_status_date,
                "product_no": row.product_no,
                "drug_name": row.drug_name,
                "active_ingredient": row.active_ingredient,
                "marketing_status": row.marketing_status,
            })

        if i % 10 == 0:
            log(f"  gate 3/enrichment processed: {i}/{len(universe)}")

    wb = Workbook()

    ws = wb.active
    ws.title = "CompanySummary"
    write_sheet(ws, summary_rows if summary_rows else [{"message": "No company summary rows"}])

    ws2 = wb.create_sheet("Universe")
    write_sheet(ws2, universe_rows if universe_rows else [{"message": "No universe rows"}])

    ws3 = wb.create_sheet("TrialsRaw")
    write_sheet(ws3, trials_raw_rows if trials_raw_rows else [{"message": "No trials matched"}])

    ws4 = wb.create_sheet("FDARaw")
    write_sheet(ws4, fda_raw_rows if fda_raw_rows else [{"message": "No FDA rows matched"}])

    ws5 = wb.create_sheet("Unmatched")
    write_sheet(ws5, unmatched if unmatched else [{"message": "No unmatched companies"}])

    ws6 = wb.create_sheet("Errors")
    write_sheet(ws6, errors if errors else [{"message": "No errors"}])

    out_path = Path(args.output).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log(f"Saved workbook: {out_path}")


def build_parser()() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Excel-first biotech screener MVP.")
    p.add_argument("--output", default="data/biotech_screener_v1.xlsx", help="Output .xlsx path")
    p.add_argument("--tickers", default="", help="Comma-separated ticker whitelist, e.g. MRNA,CRSP,VRTX")
    p.add_argument("--limit", type=int, default=0, help="Limit the SEC universe before filtering")
    p.add_argument("--only-biotech", action="store_true", help="Keep only companies that pass the biotech heuristic")
    p.add_argument("--max-trials-per-alias", type=int, default=100, help="Max CT.gov trials to fetch per alias")
    p.add_argument("--include-fda", action="store_true", help="Include Drugs@FDA sponsor matching")
    p.add_argument("--alias-overrides", default="", help="CSV file with manual ticker->alias overrides")
    return p


if __name__ == "__main__":
    parser = build_parser()
    args = parser.parse_args()
    try:
        run(args)
    except KeyboardInterrupt:
        sys.exit(130)
    except Exception:
        traceback.print_exc()
        sys.exit(1)
